#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
VPC Order Processor v4.8 - PRODUCCION
VIDIMPORT S.A. de C.V.
Listo para Render / PythonAnywhere / cualquier hosting
"""

import os
import re
import json
import logging
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple, Any
from io import BytesIO

from flask import Flask, request, jsonify, send_file, session
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font as ExcelFont, Alignment, Border, Side

# CONFIGURACION - Rutas absolutas para produccion
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, "uploads")
OUTPUT_FOLDER = os.path.join(BASE_DIR, "outputs")

app = Flask(__name__)
app.secret_key = 'vpc-secret-key-2026-vidimport'
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024  # 50MB

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

logging.basicConfig(level=logging.INFO, format='%(asctime)s | %(levelname)-8s | %(message)s')
logger = logging.getLogger('VPC-Web')

EXCEL_COLORS = {
    'vpc': 'C6EFCE', 'vpc_font': '006100',
    'otros': 'FFEB9C', 'otros_font': '9C5700',
    'sin_clasificar': 'FFC7CE', 'sin_clasificar_font': '9C0006',
    'header': '4472C4', 'header_font': 'FFFFFF',
}

# =============================================================================
# HTML PAGE
# =============================================================================

HTML_PAGE = """
<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>VPC Order Processor v4.8</title>
<style>
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:'Segoe UI',system-ui,sans-serif;background:linear-gradient(135deg,#1a1a2e 0%,#16213e 100%);color:#fff;min-height:100vh}
.container{max-width:1100px;margin:0 auto;padding:20px}
header{text-align:center;padding:30px 0;border-bottom:2px solid #e94560;margin-bottom:30px}
header h1{font-size:2.2rem}header p{color:#a0a0a0;font-size:.9rem}
.card{background:rgba(22,33,62,.8);border-radius:12px;padding:25px;margin-bottom:20px;border:1px solid rgba(255,255,255,.1)}
.card-title{font-size:1.1rem;font-weight:600;margin-bottom:15px}
.modo-selector{display:flex;gap:15px;margin-bottom:20px}
.modo-option{flex:1;padding:15px;border-radius:10px;cursor:pointer;text-align:center;border:2px solid transparent;transition:all .3s}
.modo-option:hover{transform:translateY(-2px)}
.modo-option.city{background:linear-gradient(135deg,rgba(232,93,4,.2),rgba(250,163,7,.1));border-color:#E85D04}
.modo-option.city.sel{background:linear-gradient(135deg,rgba(232,93,4,.4),rgba(250,163,7,.2));box-shadow:0 0 15px rgba(232,93,4,.3)}
.modo-option.fresko{background:linear-gradient(135deg,rgba(45,106,79,.2),rgba(82,183,136,.1));border-color:#2D6A4F}
.modo-option.fresko.sel{background:linear-gradient(135deg,rgba(45,106,79,.4),rgba(82,183,136,.2));box-shadow:0 0 15px rgba(45,106,79,.3)}
.modo-option .icon{font-size:1.8rem;margin-bottom:5px}
.modo-option .label{font-size:1rem;font-weight:600}
.file-group{margin-bottom:15px}
.file-group label{display:block;margin-bottom:6px;font-weight:500;color:#74b9ff}
.dropzone{border:2px dashed rgba(255,255,255,.3);border-radius:8px;padding:25px;text-align:center;cursor:pointer;transition:all .3s;position:relative}
.dropzone:hover{border-color:#e94560;background:rgba(233,69,96,.05)}
.dropzone input{position:absolute;top:0;left:0;width:100%;height:100%;opacity:0;cursor:pointer}
.dropzone .icon{font-size:2rem;margin-bottom:8px}
.dropzone .text{color:#a0a0a0;font-size:.9rem}
.dropzone .fn{color:#00b894;font-weight:600;margin-top:6px;display:none}
.dropzone .fn.show{display:block}
.btn-procesar{width:100%;padding:16px;font-size:1.1rem;font-weight:700;border:none;border-radius:10px;cursor:pointer;transition:all .3s;text-transform:uppercase;letter-spacing:1px;color:#fff}
.btn-procesar.city{background:linear-gradient(135deg,#E85D04,#FAA307)}
.btn-procesar.fresko{background:linear-gradient(135deg,#2D6A4F,#52B788)}
.btn-procesar:hover{transform:translateY(-2px);box-shadow:0 8px 25px rgba(0,0,0,.3)}
.btn-procesar:disabled{opacity:.6;cursor:not-allowed;transform:none}
.log-container{background:#0a0a1a;border-radius:8px;padding:12px;max-height:400px;overflow-y:auto;font-family:Consolas,monospace;font-size:.8rem}
.log-line{padding:2px 0;border-bottom:1px solid rgba(255,255,255,.05)}
.log-line .t{color:#636e72;margin-right:8px}
.log-line .lv{font-weight:bold;margin-right:6px}
.log-line .lv.i{color:#74b9ff}
.log-line .lv.w{color:#fdcb6e}
.log-line .lv.e{color:#d63031}
.log-line .lv.d{color:#b2bec3}
.log-line .m{color:#dfe6e9}
.stats{display:grid;grid-template-columns:repeat(auto-fit,minmax(120px,1fr));gap:10px;margin:15px 0}
.stat-card{background:rgba(255,255,255,.05);border-radius:8px;padding:12px;text-align:center}
.stat-card .n{font-size:1.8rem;font-weight:700}
.stat-card .l{font-size:.8rem;color:#a0a0a0}
.stat-card.tot .n{color:#74b9ff}
.stat-card.vpc .n{color:#00b894}
.stat-card.otr .n{color:#fdcb6e}
.stat-card.sin .n{color:#d63031}
.btn-dl{display:inline-block;padding:12px 30px;background:linear-gradient(135deg,#00b894,#00cec9);color:#fff;text-decoration:none;border-radius:8px;font-weight:700;transition:all .3s}
.btn-dl:hover{transform:translateY(-2px);box-shadow:0 8px 20px rgba(0,184,148,.3)}
.alert{padding:12px;border-radius:8px;margin:10px 0;font-size:.9rem}
.alert.err{background:rgba(214,48,49,.2);border:1px solid #d63031;color:#fab1a0}
.alert.ok{background:rgba(0,184,148,.2);border:1px solid #00b894;color:#55efc4}
.hidden{display:none !important}
.spinner{display:inline-block;width:18px;height:18px;border:3px solid rgba(255,255,255,.3);border-radius:50%;border-top-color:#fff;animation:spin 1s linear infinite;margin-right:8px}
@keyframes spin{to{transform:rotate(360deg)}}
.resumen-table{width:100%;border-collapse:collapse;font-size:.85rem;margin-top:10px}
.resumen-table th{background:rgba(68,114,196,.4);padding:10px;text-align:center;font-weight:600;border:1px solid rgba(255,255,255,.2)}
.resumen-table td{padding:8px 10px;border:1px solid rgba(255,255,255,.1);text-align:center}
.resumen-table tr:hover{background:rgba(255,255,255,.03)}
.resumen-table tr.total-row{font-weight:bold;background:rgba(255,255,255,.08)}
.resumen-table .col-tienda{text-align:left;font-weight:600}
.resumen-table .col-num{font-weight:700}
@media(max-width:768px){.modo-selector{flex-direction:column}header h1{font-size:1.6rem}}
</style>
</head>
<body>
<div class="container">
<header><h1>VPC Order Processor v4.8</h1><p>VIDIMPORT S.A. de C.V. | Procesamiento de pedidos</p></header>

<div class="card">
<div class="card-title">Selecciona la cadena</div>
<div class="modo-selector">
<div class="modo-option city sel" onclick="selModo('city')">
<div class="icon">&#127978;</div><div class="label">City Market</div>
</div>
<div class="modo-option fresko" onclick="selModo('fresko')">
<div class="icon">&#129388;</div><div class="label">Fresko</div>
</div>
</div>
</div>

<div class="card">
<div class="card-title">Archivos de entrada</div>
<div class="file-group">
<label>Pedidos del dia (Excel)</label>
<div class="dropzone" onclick="document.getElementById('fp').click()">
<input type="file" id="fp" accept=".xlsx,.xls" onchange="hf(this,'p')">
<div class="icon">&#128202;</div>
<div class="text">Arrastra aqui o haz clic<br><small>.xlsx o .xls</small></div>
<div class="fn" id="fn-p"></div>
</div>
</div>
<div class="file-group">
<label>Catalogo SAP (.xlsm o .xlsx)</label>
<div class="dropzone" onclick="document.getElementById('fc').click()">
<input type="file" id="fc" accept=".xlsx,.xls,.xlsm" onchange="hf(this,'c')">
<div class="icon">&#128209;</div>
<div class="text">Arrastra aqui o haz clic<br><small>.xlsx, .xls o .xlsm</small></div>
<div class="fn" id="fn-c"></div>
</div>
</div>
<button class="btn-procesar city" id="btn" onclick="proc()">&#128640; PROCESAR PEDIDOS</button>
</div>

<div class="card">
<div class="card-title">&#128203; Log de procesamiento</div>
<div class="log-container" id="log">
<div class="log-line"><span class="t">--:--:--</span><span class="lv i">INFO</span><span class="m">Esperando archivos...</span></div>
</div>
</div>

<div class="card hidden" id="res">
<div class="card-title">&#9989; Resultados</div>
<div class="stats" id="stats"></div>
<div id="alerts"></div>
<div style="text-align:center;margin:15px 0">
<a href="/api/descargar" class="btn-dl" id="bdl" onclick="return dl()">&#11015; DESCARGAR EXCEL</a>
</div>
<div>
<div class="card-title">Resumen de Cajas por Tienda</div>
<div style="overflow-x:auto"><table class="resumen-table" id="tbl_resumen"><thead><tr></tr></thead><tbody></tbody></table></div>
</div>
</div>
</div>

<script>
let modo='city',files={p:null,c:null};
function selModo(m){modo=m;document.querySelectorAll('.modo-option').forEach(e=>e.classList.remove('sel'));document.querySelector('.modo-option.'+m).classList.add('sel');document.getElementById('btn').className='btn-procesar '+m;}
function hf(i,t){if(i.files&&i.files[0]){files[t]=i.files[0];document.getElementById('fn-'+t).textContent='OK '+i.files[0].name;document.getElementById('fn-'+t).classList.add('show');lg('INFO','Archivo: '+i.files[0].name+' ('+(i.files[0].size/1024/1024).toFixed(2)+' MB)');}}
function lg(l,m){const c=document.getElementById('log'),n=new Date().toLocaleTimeString('es-MX');const d=document.createElement('div');d.className='log-line';d.innerHTML='<span class="t">'+n+'</span><span class="lv '+l.toLowerCase().substring(0,1)+'">'+l+'</span><span class="m">'+m+'</span>';c.appendChild(d);c.scrollTop=c.scrollHeight;}
function proc(){if(!files.p||!files.c){alert('Selecciona ambos archivos');return;}const b=document.getElementById('btn');b.disabled=true;b.innerHTML='<span class="spinner"></span> Procesando...';document.getElementById('res').classList.add('hidden');document.getElementById('log').innerHTML='';lg('INFO','Iniciando...');const fd=new FormData();fd.append('pedidos',files.p);fd.append('catalogo',files.c);fd.append('modo',modo);fetch('/api/procesar',{method:'POST',body:fd}).then(r=>{if(!r.ok)throw new Error('HTTP '+r.status);return r.json();}).then(d=>{b.disabled=false;b.innerHTML='&#128640; PROCESAR PEDIDOS';if(d.logs)d.logs.forEach(l=>lg(l.level,l.message));if(d.success){showRes(d);}else{lg('ERROR',d.error||'Error desconocido');document.getElementById('res').classList.remove('hidden');}}).catch(e=>{b.disabled=false;b.innerHTML='&#128640; PROCESAR PEDIDOS';lg('ERROR','Error: '+e.message);});}
function showRes(d){const r=document.getElementById('res');r.classList.remove('hidden');const s=d.stats;const grid=document.getElementById('stats');grid.innerHTML='<div class="stat-card tot"><div class="n">'+s.total+'</div><div class="l">Total Renglones</div></div><div class="stat-card vpc"><div class="n">'+s.vpc+'</div><div class="l">Producto VPC</div></div><div class="stat-card otr"><div class="n">'+s.otros+'</div><div class="l">Otros Prov.</div></div><div class="stat-card sin"><div class="n">'+s.sin_clasificar+'</div><div class="l">Sin Clasif.</div></div><div class="stat-card tot"><div class="n">'+s.tiendas+'</div><div class="l">Tiendas</div></div>';const a=document.getElementById('alerts');if(s.sin_clasificar>0){a.innerHTML='<div class="alert err">&#9888; Hay '+s.sin_clasificar+' productos sin clasificar. Revisa la pestana SIN CLASIFICAR en el Excel.</div>';}else{a.innerHTML='<div class="alert ok">&#9989; Todos los productos clasificados correctamente.</div>';}if(s.resumen_cajas&&s.resumen_cajas.length>0){const t=document.getElementById('tbl_resumen');const headers=['TIENDA','NOMBRE TIENDA','PRODUCTO VPC','OTROS PROVEEDORES','SIN CLASIFICAR','TOTAL'];t.querySelector('thead tr').innerHTML=headers.map(x=>'<th>'+x+'</th>').join('');t.querySelector('tbody').innerHTML=s.resumen_cajas.map((row,idx)=>{const isTotal=row.tienda==='TOTAL';const cls=isTotal?'total-row':'';return'<tr class="'+cls+'"><td class="col-tienda">'+row.tienda+'</td><td>'+row.nombre+'</td><td class="col-num">'+row.vpc+'</td><td class="col-num">'+row.otros+'</td><td class="col-num">'+row.sin_clasificar+'</td><td class="col-num">'+row.total+'</td></tr>';}).join('');}r.scrollIntoView({behavior:'smooth'});}
function dl(){const b=document.getElementById('bdl');b.textContent='&#11015; Descargando...';setTimeout(()=>b.textContent='&#11015; DESCARGAR EXCEL',3000);return true;}
['p','c'].forEach(t=>{const z=document.getElementById('fp').parentElement;if(t==='c')z=document.getElementById('fc').parentElement;z.addEventListener('dragover',e=>{e.preventDefault();z.style.borderColor='#e94560';z.style.background='rgba(233,69,96,.05)';});z.addEventListener('dragleave',()=>{z.style.borderColor='rgba(255,255,255,.3)';z.style.background='transparent';});z.addEventListener('drop',e=>{e.preventDefault();z.style.borderColor='rgba(255,255,255,.3)';z.style.background='transparent';const i=document.getElementById('f'+t);i.files=e.dataTransfer.files;hf(i,t);});});
</script>
</body>
</html>
"""

# =============================================================================
# PROCESADOR
# =============================================================================

class OrderProcessor:
    def __init__(self):
        self.catalogo_sap = {}
        self.catalogo_otros = {}
        self.log_messages = []

    def _log(self, level: str, msg: str):
        self.log_messages.append({'time': datetime.now().strftime('%H:%M:%S'), 'level': level, 'message': msg})
        logger.info(f"[{level}] {msg}")

    def _limpiar_codigo(self, codigo) -> str:
        if codigo is None: return ""
        codigo = str(codigo).strip()
        codigo = re.sub(r'\s+', '', codigo)
        try:
            if '.' in codigo:
                num = float(codigo)
                if num == int(num): codigo = str(int(num))
        except: pass
        return codigo

    def _limpiar_texto(self, texto) -> str:
        if texto is None: return ""
        return str(texto).strip()

    def _extraer_numero_cajas(self, valor) -> int:
        if valor is None or pd.isna(valor): return 0
        texto = str(valor).strip().lower()
        if 'kg' in texto:
            match = re.search(r'(\d+)', texto)
            return int(match.group(1)) if match else 0
        try: return int(float(texto))
        except: return 0

    def _es_codigo_valido(self, codigo) -> bool:
        if not codigo: return False
        codigo_limpio = self._limpiar_codigo(codigo)
        return bool(codigo_limpio) and codigo_limpio[0].isdigit()

    def _es_descripcion_valida(self, texto) -> bool:
        if texto is None or pd.isna(texto):
            return False
        texto_str = str(texto).strip()
        if len(texto_str) < 3:
            return False
        if any(x in texto_str.upper() for x in ['TOTAL', 'CAJAS', 'PROMEDIO', 'CODIGO', 'CÓDIGO', 'DESCRIPCION', 'DESCRIPCIÓN']):
            return False
        if not any(c.isalpha() for c in texto_str):
            return False
        return True

    def cargar_catalogo_sap(self, ruta_archivo: str, pestaña_vpc: str) -> bool:
        try:
            self._log('INFO', f'Cargando catalogo SAP: {pestaña_vpc}')
            engine = 'openpyxl' if ruta_archivo.endswith('.xlsm') else None
            # Optimizacion: leer solo columnas necesarias (C a H = cols 2-7)
            df = pd.read_excel(
                ruta_archivo, 
                sheet_name=pestaña_vpc, 
                header=0, 
                engine=engine,
                usecols='C:H',
                dtype=str
            )
            self._log('INFO', f'Filas en catalogo SAP: {len(df)}')
            self.catalogo_sap = {}
            for idx, row in df.iterrows():
                try:
                    codigo = self._limpiar_codigo(row.iloc[0] if len(row) > 0 else None)
                    if not codigo: continue
                    self.catalogo_sap[codigo] = {
                        'descripcion_cliente': self._limpiar_texto(row.iloc[1] if len(row) > 1 else ''),
                        'descripcion_sap': self._limpiar_texto(row.iloc[2] if len(row) > 2 else ''),
                        'material_sap': self._limpiar_texto(row.iloc[3] if len(row) > 3 else ''),
                        'cajas': self._extraer_numero_cajas(row.iloc[4] if len(row) > 4 else 0),
                        'precio': row.iloc[5] if len(row) > 5 else 0,
                    }
                except Exception as e:
                    self._log('WARNING', f'Error fila {idx} catalogo SAP: {e}')
                    continue
            self._log('INFO', f'Catalogo SAP cargado: {len(self.catalogo_sap)} productos')
            return True
        except Exception as e:
            self._log('ERROR', f'Error cargando catalogo SAP: {e}')
            return False

    def cargar_catalogo_otros(self, ruta_archivo: str, pestaña_otros: str, col_producto: str) -> bool:
        try:
            self._log('INFO', f'Cargando catalogo otros: {pestaña_otros}')
            engine = 'openpyxl' if ruta_archivo.endswith('.xlsm') else None
            # Optimizacion: leer solo columnas A y C (codigo y producto)
            df = pd.read_excel(
                ruta_archivo, 
                sheet_name=pestaña_otros, 
                header=0, 
                engine=engine,
                usecols='A,C',
                dtype=str
            )
            self._log('INFO', f'Filas en catalogo otros: {len(df)}')
            self.catalogo_otros = {}
            for idx, row in df.iterrows():
                try:
                    codigo = self._limpiar_codigo(row.iloc[0] if len(row) > 0 else None)
                    producto = self._limpiar_texto(row.iloc[1] if len(row) > 1 else '')
                    if codigo: self.catalogo_otros[codigo] = producto
                except Exception as e:
                    self._log('WARNING', f'Error fila {idx} catalogo otros: {e}')
                    continue
            self._log('INFO', f'Catalogo otros cargado: {len(self.catalogo_otros)} productos')
            return True
        except Exception as e:
            self._log('ERROR', f'Error cargando catalogo otros: {e}')
            return False

    def clasificar_producto(self, codigo: str) -> Tuple[str, Dict]:
        codigo_limpio = self._limpiar_codigo(codigo)
        if not codigo_limpio: return "Sin Clasificar", {"estatus": "Codigo vacio"}
        if codigo_limpio in self.catalogo_sap:
            datos = self.catalogo_sap[codigo_limpio]
            return "Producto VPC", {
                'descripcion_cliente': datos['descripcion_cliente'],
                'descripcion_sap': datos['descripcion_sap'],
                'material_sap': datos['material_sap'],
                'cajas_sap': 0,
                'precio': datos['precio'],
            }
        if codigo_limpio in self.catalogo_otros:
            return "Otros Proveedores", {'producto': self.catalogo_otros[codigo_limpio]}
        return "Sin Clasificar", {"estatus": "No se encontro en catalogo ni en inventario - Revisar"}

    def procesar_city_market(self, ruta_pedidos: str) -> List[Dict]:
        self._log('INFO', '=== PROCESANDO CITY MARKET ===')
        resultados = []
        try:
            xls = pd.ExcelFile(ruta_pedidos)
            pestañas = xls.sheet_names
            self._log('INFO', f'Pestañas encontradas: {pestañas}')
            for pestaña in pestañas:
                if not re.match(r'^\d+$', str(pestaña).strip()):
                    self._log('INFO', f'Ignorando pestaña no numerica: {pestaña}')
                    continue
                num_tienda = str(pestaña).strip()
                self._log('INFO', f'Procesando tienda: {num_tienda}')
                df = pd.read_excel(ruta_pedidos, sheet_name=pestaña, header=None)
                if df.empty or len(df) < 9:
                    self._log('WARNING', f'Hoja {pestaña} muy corta, saltando')
                    continue
                nombre_tienda = ""
                try:
                    nombre_celda = df.iloc[2, 11] if df.shape[1] > 11 else None
                    if nombre_celda and not pd.isna(nombre_celda):
                        nombre_tienda = self._limpiar_texto(nombre_celda)
                        match = re.search(r'SUCURSAL Nº :\s*\d+\s*(.+)', nombre_tienda)
                        if match: nombre_tienda = match.group(1).strip()
                except: pass
                if not nombre_tienda: nombre_tienda = f"Tienda {num_tienda}"
                self._log('INFO', f'Nombre tienda: {nombre_tienda}')
                bloques = [(1, 2, 3), (7, 8, 9), (13, 14, 15)]
                filas_procesadas = 0
                for idx in range(8, len(df)):
                    fila = df.iloc[idx]
                    for col_cod, col_desc, col_cajas in bloques:
                        try:
                            if col_cod >= len(fila) or col_desc >= len(fila): continue
                            codigo = fila.iloc[col_cod]
                            descripcion = fila.iloc[col_desc] if col_desc < len(fila) else None
                            cajas = fila.iloc[col_cajas] if col_cajas < len(fila) else 0
                            if not self._es_codigo_valido(codigo): continue
                            codigo_limpio = self._limpiar_codigo(codigo)
                            desc_limpia = self._limpiar_texto(descripcion)
                            num_cajas = self._extraer_numero_cajas(cajas)
                            if not desc_limpia: continue
                            tipo, datos_extra = self.clasificar_producto(codigo_limpio)
                            registro = {
                                'num_tienda': num_tienda, 'nombre_tienda': nombre_tienda,
                                'codigo': codigo_limpio, 'descripcion_pedido': desc_limpia,
                                'cajas_pedido': num_cajas, 'tipo': tipo, **datos_extra,
                            }
                            if tipo == 'Producto VPC':
                                registro['cajas_sap'] = num_cajas
                            resultados.append(registro)
                            filas_procesadas += 1
                        except: continue
                self._log('INFO', f'Filas procesadas tienda {num_tienda}: {filas_procesadas}')
            self._log('INFO', f'TOTAL City Market: {len(resultados)} renglones')
            return resultados
        except Exception as e:
            self._log('ERROR', f'Error procesando City Market: {e}')
            import traceback
            self._log('ERROR', traceback.format_exc())
            return []

    # =========================================================================
    # FRESKO - VERSION v4.8 (openpyxl + deteccion automatica + hojas ocultas)
    # =========================================================================
    def _detectar_bloques_fresko(self, ws) -> List[Tuple[int, int, int]]:
        max_row = ws.max_row
        max_col = ws.max_column

        header_row = None
        for r in range(1, min(15, max_row + 1)):
            for c in range(1, max_col + 1):
                cell_val = ws.cell(row=r, column=c).value
                if cell_val and isinstance(cell_val, str):
                    val_upper = cell_val.upper().strip()
                    if "CODIGO" in val_upper or "CÓDIGO" in val_upper:
                        for c2 in range(c + 1, min(c + 5, max_col + 1)):
                            val2 = ws.cell(row=r, column=c2).value
                            if val2 and isinstance(val2, str) and ("DESCRIPCION" in val2.upper() or "DESCRIPCIÓN" in val2.upper()):
                                header_row = r
                                break
                if header_row:
                    break
            if header_row:
                break

        if not header_row:
            self._log('WARNING', 'No se encontro fila de encabezados con CODIGO y DESCRIPCION')
            return []

        self._log('INFO', f'Fila de encabezados detectada: {header_row}')

        posiciones = []
        for c in range(1, max_col + 1):
            val = ws.cell(row=header_row, column=c).value
            if val and isinstance(val, str):
                v = val.upper().strip()
                if "CODIGO" in v or "CÓDIGO" in v:
                    posiciones.append((c, 'CODIGO'))
                elif "DESCRIPCION" in v or "DESCRIPCIÓN" in v:
                    posiciones.append((c, 'DESCRIPCION'))
                elif "PRECIO" in v:
                    posiciones.append((c, 'PRECIO'))

        posiciones.sort(key=lambda x: x[0])

        bloques = []
        i = 0
        while i < len(posiciones):
            if posiciones[i][1] == 'CODIGO':
                col_cod = posiciones[i][0]
                col_desc = None
                col_precio = None

                for j in range(i + 1, len(posiciones)):
                    if posiciones[j][1] == 'DESCRIPCION' and col_desc is None:
                        col_desc = posiciones[j][0]
                    if posiciones[j][1] == 'PRECIO' and col_precio is None:
                        col_precio = posiciones[j][0]
                    if col_desc and col_precio:
                        break

                if col_desc and col_precio:
                    cajas_col = None
                    for test_row in range(header_row + 2, min(header_row + 10, max_row + 1)):
                        cod_val = ws.cell(row=test_row, column=col_cod).value
                        if cod_val and str(cod_val).strip() and str(cod_val).strip()[0].isdigit():
                            for c in range(col_desc + 1, col_precio):
                                val = ws.cell(row=test_row, column=c).value
                                if val is not None and val != '':
                                    try:
                                        v = float(str(val).strip())
                                        if 0 <= v <= 500:
                                            cajas_col = c
                                            break
                                    except:
                                        pass
                            if cajas_col:
                                break

                    if cajas_col is None:
                        cajas_col = col_desc + 1

                    bloques.append((col_cod, col_desc, cajas_col))
                    self._log('INFO', f'Bloque detectado: COD={col_cod}, DESC={col_desc}, CAJAS={cajas_col}')
            i += 1

        return bloques

    def _extraer_fecha_entrega_fresko(self, ws) -> str:
        max_col = ws.max_column

        for r in range(1, 10):
            for c in range(1, max_col + 1):
                val = ws.cell(row=r, column=c).value
                if val and isinstance(val, str):
                    if "FECHA" in val.upper() and "ENTREGA" in val.upper():
                        for c2 in range(c + 1, min(c + 5, max_col + 1)):
                            val2 = ws.cell(row=r, column=c2).value
                            if val2 and isinstance(val2, datetime):
                                return val2.strftime("%Y-%m-%d")
                            if val2 and isinstance(val2, str):
                                m = re.search(r'(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})', val2)
                                if m:
                                    d, mth, y = m.groups()
                                    y = int(y)
                                    if y < 100:
                                        y += 2000
                                    return f"{y:04d}-{int(mth):02d}-{int(d):02d}"

        for r in range(1, 10):
            for c in range(1, max_col + 1):
                val = ws.cell(row=r, column=c).value
                if val and isinstance(val, datetime):
                    return val.strftime("%Y-%m-%d")
                if val and isinstance(val, str):
                    m = re.search(r'(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})', val)
                    if m:
                        d, mth, y = m.groups()
                        y = int(y)
                        if y < 100:
                            y += 2000
                        return f"{y:04d}-{int(mth):02d}-{int(d):02d}"

        return ""

    def procesar_fresko(self, ruta_pedidos: str) -> List[Dict]:
        self._log('INFO', '=== PROCESANDO FRESKO v4.8 (openpyxl + deteccion automatica + solo hojas visibles) ===')
        resultados = []

        try:
            wb = openpyxl.load_workbook(ruta_pedidos, data_only=True)
            hojas = wb.sheetnames

            # Filtrar solo hojas VISIBLES que son numeros de tienda
            hojas_tienda = []
            for h in hojas:
                if not h.strip().isdigit():
                    continue
                ws_check = wb[h]
                if ws_check.sheet_state == 'hidden':
                    self._log('INFO', f'Ignorando hoja OCULTA: {h}')
                    continue
                hojas_tienda.append(h)

            self._log('INFO', f'Hojas de tienda VISIBLES detectadas: {hojas_tienda}')

            for hoja in hojas_tienda:
                num_tienda = str(hoja).strip()
                self._log('INFO', f'Procesando tienda: {num_tienda}')

                ws = wb[hoja]

                bloques = self._detectar_bloques_fresko(ws)
                if not bloques:
                    self._log('WARNING', f'Tienda {num_tienda}: No se detectaron bloques de datos')
                    continue

                fecha = self._extraer_fecha_entrega_fresko(ws)
                if fecha:
                    self._log('INFO', f'Fecha de entrega: {fecha}')

                max_row = ws.max_row
                filas_tienda = 0
                cajas_tienda = 0

                for bloque_idx, (col_cod, col_desc, col_cajas) in enumerate(bloques):
                    filas_bloque = 0
                    cajas_bloque = 0

                    for r in range(1, max_row + 1):
                        codigo = ws.cell(row=r, column=col_cod).value
                        if not codigo or str(codigo).strip() == '':
                            continue
                        cod_str = str(codigo).strip()
                        if not cod_str[0].isdigit():
                            continue

                        descripcion = ws.cell(row=r, column=col_desc).value
                        if not descripcion:
                            continue
                        desc_str = str(descripcion).strip()
                        if not desc_str or desc_str.upper() in ['TOTAL CAJAS', 'CODIGO', 'CÓDIGO', 'DESCRIPCION', 'DESCRIPCIÓN']:
                            continue

                        cajas_val = ws.cell(row=r, column=col_cajas).value
                        num_cajas = 0
                        if cajas_val is not None and cajas_val != '':
                            try:
                                num_cajas = int(float(str(cajas_val).strip()))
                            except:
                                pass

                        if num_cajas <= 0:
                            continue

                        codigo_limpio = self._limpiar_codigo(cod_str)
                        tipo, datos_extra = self.clasificar_producto(codigo_limpio)

                        registro = {
                            'num_tienda': num_tienda,
                            'nombre_tienda': num_tienda,
                            'codigo': codigo_limpio,
                            'descripcion_pedido': desc_str,
                            'cajas_pedido': num_cajas,
                            'tipo': tipo,
                            **datos_extra,
                        }
                        if tipo == 'Producto VPC':
                            registro['cajas_sap'] = num_cajas

                        resultados.append(registro)
                        filas_bloque += 1
                        cajas_bloque += num_cajas
                        filas_tienda += 1
                        cajas_tienda += num_cajas

                    self._log('INFO', f'  Bloque {bloque_idx+1}: {filas_bloque} articulos, {cajas_bloque} cajas')

                self._log('INFO', f'Tienda {num_tienda} TOTAL: {filas_tienda} filas, {cajas_tienda} cajas')

            self._log('INFO', f'TOTAL Fresko: {len(resultados)} renglones')
            return resultados

        except Exception as e:
            self._log('ERROR', f'Error procesando Fresko: {e}')
            import traceback
            self._log('ERROR', traceback.format_exc())
            return []

    def generar_excel(self, resultados: List[Dict], modo: str) -> BytesIO:
        try:
            self._log('INFO', 'Generando Excel...')
            if not resultados:
                self._log('WARNING', 'No hay resultados para exportar')
                return None
            df = pd.DataFrame(resultados)
            col_producto = 'PRODUCTO ACCES' if modo == 'city' else 'PRODUCTO'
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            header_fill = PatternFill(start_color=EXCEL_COLORS['header'], end_color=EXCEL_COLORS['header'], fill_type='solid')
            header_font = ExcelFont(color=EXCEL_COLORS['header_font'], bold=True, size=11)
            vpc_fill = PatternFill(start_color=EXCEL_COLORS['vpc'], end_color=EXCEL_COLORS['vpc'], fill_type='solid')
            vpc_font = ExcelFont(color=EXCEL_COLORS['vpc_font'])
            otros_fill = PatternFill(start_color=EXCEL_COLORS['otros'], end_color=EXCEL_COLORS['otros'], fill_type='solid')
            otros_font = ExcelFont(color=EXCEL_COLORS['otros_font'])
            sin_fill = PatternFill(start_color=EXCEL_COLORS['sin_clasificar'], end_color=EXCEL_COLORS['sin_clasificar'], fill_type='solid')
            sin_font = ExcelFont(color=EXCEL_COLORS['sin_clasificar_font'])
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            def escribir_hoja(ws, df_sub, headers, cols, fill, font, border, hfill, hfont):
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.fill = hfill; cell.font = hfont; cell.alignment = Alignment(horizontal='center'); cell.border = border
                for row_idx, registro in enumerate(df_sub.to_dict('records'), 2):
                    for col_idx, col_name in enumerate(cols, 1):
                        valor = registro.get(col_name, '')
                        if valor is None or pd.isna(valor): valor = ''
                        cell = ws.cell(row=row_idx, column=col_idx, value=valor)
                        cell.border = border; cell.fill = fill; cell.font = font
                for col in ws.columns:
                    max_length = 0; column = col[0].column_letter
                    for cell in col:
                        try:
                            if cell.value: max_length = max(max_length, len(str(cell.value)))
                        except: pass
                    ws.column_dimensions[column].width = min(max_length + 2, 50)

            def escribir_hoja_simple(ws, df_sub, headers, cols, fill, font, border, hfill, hfont):
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.fill = hfill; cell.font = hfont; cell.alignment = Alignment(horizontal='center'); cell.border = border
                for row_idx, registro in enumerate(df_sub.to_dict('records'), 2):
                    for col_idx, col_name in enumerate(cols, 1):
                        valor = registro.get(col_name, '')
                        if valor is None or pd.isna(valor): valor = ''
                        cell = ws.cell(row=row_idx, column=col_idx, value=valor)
                        cell.border = border; cell.fill = fill; cell.font = font
                for col in ws.columns:
                    max_length = 0; column = col[0].column_letter
                    for cell in col:
                        try:
                            if cell.value: max_length = max(max_length, len(str(cell.value)))
                        except: pass
                    ws.column_dimensions[column].width = min(max_length + 2, 50)

            # PESTAÑA 1: RESUMEN CONSOLIDADO
            ws1 = wb.create_sheet("RESUMEN CONSOLIDADO")
            cols_consolidado = ['num_tienda', 'nombre_tienda', 'codigo', 'descripcion_pedido', 'cajas_pedido', 'tipo', 'descripcion_cliente', 'descripcion_sap', 'material_sap', 'cajas_sap', 'precio', 'producto']
            headers = ['# TIENDA', 'NOMBRE TIENDA', 'CODIGO', 'DESCRIPCION PEDIDO', 'CAJAS PEDIDO', 'TIPO', 'DESCRIPCION CLIENTE', 'DESCRIPCION SAP', 'MATERIAL SAP', 'CAJAS SAP', 'PRECIO', col_producto]
            for col_idx, header in enumerate(headers, 1):
                cell = ws1.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill; cell.font = header_font; cell.alignment = Alignment(horizontal='center'); cell.border = thin_border
            for row_idx, registro in enumerate(resultados, 2):
                for col_idx, col_name in enumerate(cols_consolidado, 1):
                    valor = registro.get(col_name, '')
                    if valor is None or pd.isna(valor): valor = ''
                    cell = ws1.cell(row=row_idx, column=col_idx, value=valor)
                    cell.border = thin_border
                    tipo = registro.get('tipo', '')
                    if tipo == 'Producto VPC': cell.fill = vpc_fill; cell.font = vpc_font
                    elif tipo == 'Otros Proveedores': cell.fill = otros_fill; cell.font = otros_font
                    elif tipo == 'Sin Clasificar': cell.fill = sin_fill; cell.font = sin_font
            for col in ws1.columns:
                max_length = 0; column = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value: max_length = max(max_length, len(str(cell.value)))
                    except: pass
                ws1.column_dimensions[column].width = min(max_length + 2, 50)

            # PESTAÑA 2: PRODUCTO VPC
            df_vpc = df[df['tipo'] == 'Producto VPC'].copy()
            if not df_vpc.empty:
                ws2 = wb.create_sheet("PRODUCTO VPC")
                escribir_hoja(ws2, df_vpc, headers, cols_consolidado, vpc_fill, vpc_font, thin_border, header_fill, header_font)

            # PESTAÑA 3: OTROS PROVEEDORES
            df_otros = df[df['tipo'] == 'Otros Proveedores'].copy()
            if not df_otros.empty:
                ws3 = wb.create_sheet("OTROS PROVEEDORES")
                cols_otros = ['num_tienda', 'nombre_tienda', 'codigo', 'descripcion_pedido', 'cajas_pedido', 'producto']
                headers_otros = ['# TIENDA', 'NOMBRE TIENDA', 'CODIGO', 'DESCRIPCION PEDIDO', 'CAJAS PEDIDO', col_producto]
                escribir_hoja_simple(ws3, df_otros, headers_otros, cols_otros, otros_fill, otros_font, thin_border, header_fill, header_font)

            # PESTAÑA 4: SIN CLASIFICAR
            df_sin = df[df['tipo'] == 'Sin Clasificar'].copy()
            if not df_sin.empty:
                ws4 = wb.create_sheet("SIN CLASIFICAR")
                cols_sin = ['num_tienda', 'nombre_tienda', 'codigo', 'descripcion_pedido', 'cajas_pedido', 'estatus']
                headers_sin = ['# TIENDA', 'NOMBRE TIENDA', 'CODIGO', 'DESCRIPCION PEDIDO', 'CAJAS PEDIDO', 'ESTATUS']
                escribir_hoja_simple(ws4, df_sin, headers_sin, cols_sin, sin_fill, sin_font, thin_border, header_fill, header_font)

            # PESTAÑA 5: RESUMEN POR TIENDA
            ws5 = wb.create_sheet("RESUMEN POR TIENDA")
            resumen = df.groupby(['num_tienda', 'nombre_tienda'])['tipo'].value_counts().unstack(fill_value=0)
            for col in ['Producto VPC', 'Otros Proveedores', 'Sin Clasificar']:
                if col not in resumen.columns: resumen[col] = 0
            resumen['TOTAL'] = resumen.sum(axis=1); resumen = resumen.reset_index()
            headers_resumen = ['# TIENDA', 'NOMBRE TIENDA', 'Producto VPC', 'Otros Proveedores', 'Sin Clasificar', 'TOTAL']
            cols_resumen = ['num_tienda', 'nombre_tienda', 'Producto VPC', 'Otros Proveedores', 'Sin Clasificar', 'TOTAL']
            for col_idx, header in enumerate(headers_resumen, 1):
                cell = ws5.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill; cell.font = header_font; cell.alignment = Alignment(horizontal='center'); cell.border = thin_border
            for row_idx, (_, row) in enumerate(resumen.iterrows(), 2):
                for col_idx, col_name in enumerate(cols_resumen, 1):
                    cell = ws5.cell(row=row_idx, column=col_idx, value=row.get(col_name, 0))
                    cell.border = thin_border; cell.alignment = Alignment(horizontal='center')
            for col in ws5.columns:
                max_length = 0; column = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value: max_length = max(max_length, len(str(cell.value)))
                    except: pass
                ws5.column_dimensions[column].width = min(max_length + 2, 30)

            output = BytesIO()
            wb.save(output)
            output.seek(0)
            self._log('INFO', 'Excel generado exitosamente')
            return output
        except Exception as e:
            self._log('ERROR', f'Error generando Excel: {e}')
            import traceback
            self._log('ERROR', traceback.format_exc())
            return None


# =============================================================================
# RUTAS FLASK
# =============================================================================

@app.route('/')
def index():
    return HTML_PAGE

@app.route('/api/procesar', methods=['POST'])
def api_procesar():
    try:
        modo = request.form.get('modo', 'city')
        logger.info(f"API procesar llamado - modo: {modo}")

        if 'pedidos' not in request.files or 'catalogo' not in request.files:
            logger.warning("Faltan archivos en la peticion")
            return jsonify({'success': False, 'error': 'Faltan archivos', 'logs': []}), 400

        pedidos_file = request.files['pedidos']
        catalogo_file = request.files['catalogo']

        if pedidos_file.filename == '' or catalogo_file.filename == '':
            return jsonify({'success': False, 'error': 'Archivos vacios', 'logs': []}), 400

        logger.info(f"Archivos recibidos: pedidos={pedidos_file.filename}, catalogo={catalogo_file.filename}")

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        temp_dir = tempfile.gettempdir()
        pedidos_path = os.path.join(temp_dir, f'pedidos_{timestamp}.xlsx')
        catalogo_path = os.path.join(temp_dir, f'catalogo_{timestamp}.xlsm')

        pedidos_file.save(pedidos_path)
        catalogo_file.save(catalogo_path)
        logger.info("Archivos guardados temporalmente")

        processor = OrderProcessor()

        if modo == 'city':
            sap_ok = processor.cargar_catalogo_sap(catalogo_path, 'SAP CITYS FRUTAMERICA')
            otros_ok = processor.cargar_catalogo_otros(catalogo_path, 'CAT CITY', 'PRODUCTO ACCES')
        else:
            sap_ok = processor.cargar_catalogo_sap(catalogo_path, 'SAP FRESKO VIDI')
            otros_ok = processor.cargar_catalogo_otros(catalogo_path, 'CAT FRESKO', 'PRODUCTO')

        if not sap_ok or not otros_ok:
            return jsonify({
                'success': False,
                'error': 'Error cargando catalogos. Verifica que el archivo SAP tenga las pestanas correctas.',
                'logs': processor.log_messages
            }), 400

        if modo == 'city':
            resultados = processor.procesar_city_market(pedidos_path)
        else:
            resultados = processor.procesar_fresko(pedidos_path)

        if not resultados:
            return jsonify({
                'success': False,
                'error': 'No se encontraron datos validos. Verifica la estructura del archivo de pedidos.',
                'logs': processor.log_messages
            }), 400

        excel_buffer = processor.generar_excel(resultados, modo)

        if not excel_buffer:
            return jsonify({
                'success': False,
                'error': 'Error generando el Excel',
                'logs': processor.log_messages
            }), 500

        temp_excel_path = os.path.join(temp_dir, f'resultado_{timestamp}.xlsx')
        with open(temp_excel_path, 'wb') as f:
            f.write(excel_buffer.getvalue())
        session['excel_path'] = temp_excel_path
        session['modo'] = modo
        session['fecha'] = datetime.now().strftime('%d%m%Y')

        df = pd.DataFrame(resultados)

        stats = {
            'total': len(resultados),
            'vpc': len(df[df['tipo'] == 'Producto VPC']),
            'otros': len(df[df['tipo'] == 'Otros Proveedores']),
            'sin_clasificar': len(df[df['tipo'] == 'Sin Clasificar']),
            'tiendas': df['num_tienda'].nunique(),
        }

        pivot_cajas = df.pivot_table(
            index=['num_tienda', 'nombre_tienda'],
            columns='tipo',
            values='cajas_pedido',
            aggfunc='sum',
            fill_value=0
        ).reset_index()

        for col in ['Producto VPC', 'Otros Proveedores', 'Sin Clasificar']:
            if col not in pivot_cajas.columns:
                pivot_cajas[col] = 0

        pivot_cajas['TOTAL'] = pivot_cajas[['Producto VPC', 'Otros Proveedores', 'Sin Clasificar']].sum(axis=1)

        total_vpc = int(pivot_cajas['Producto VPC'].sum())
        total_otros = int(pivot_cajas['Otros Proveedores'].sum())
        total_sin = int(pivot_cajas['Sin Clasificar'].sum())
        total_general = total_vpc + total_otros + total_sin

        resumen_cajas = []
        for _, row in pivot_cajas.iterrows():
            resumen_cajas.append({
                'tienda': str(row['num_tienda']),
                'nombre': str(row['nombre_tienda']),
                'vpc': int(row['Producto VPC']),
                'otros': int(row['Otros Proveedores']),
                'sin_clasificar': int(row['Sin Clasificar']),
                'total': int(row['TOTAL'])
            })

        resumen_cajas.append({
            'tienda': 'TOTAL',
            'nombre': 'TOTAL GENERAL',
            'vpc': total_vpc,
            'otros': total_otros,
            'sin_clasificar': total_sin,
            'total': total_general
        })

        stats['resumen_cajas'] = resumen_cajas

        logger.info(f"Procesamiento completado: {stats['total']} renglones, {stats['tiendas']} tiendas")

        return jsonify({
            'success': True,
            'stats': stats,
            'logs': processor.log_messages
        })

    except Exception as e:
        logger.error(f"Error en API: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return jsonify({'success': False, 'error': str(e), 'logs': []}), 500

@app.route('/api/descargar')
def api_descargar():
    try:
        excel_path = session.get('excel_path')
        modo = session.get('modo', 'city')
        fecha = session.get('fecha', datetime.now().strftime('%d%m%Y'))

        if not excel_path or not os.path.exists(excel_path):
            return jsonify({'success': False, 'error': 'No hay archivo para descargar'}), 400

        nombre = f'PEDIDOS_{"CITY_MARKET" if modo == "city" else "FRESKO"}_{fecha}.xlsx'

        return send_file(
            excel_path,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=nombre
        )

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# =============================================================================
# INICIO - Para produccion usa el PORT de la variable de entorno
# =============================================================================
if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
